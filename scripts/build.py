#!/usr/bin/env python3
"""
depremlab — ham veriyi normalize edip ilçe bazlı özet tabloyu üretir.

Girdi:  data/**  (ham indirilmiş dosyalar)
Çıktı:  derived/**  (normalize CSV'ler)

Çalıştırma:  python3 scripts/build.py
"""
import csv
import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DERIVED = ROOT / "derived"
DERIVED.mkdir(exist_ok=True)

R = 6371.0  # dünya yarıçapı, km


# --------------------------------------------------------------------------
# yardımcılar
# --------------------------------------------------------------------------
def norm_ad(s):
    """İlçe adlarını kaynaklar arası eşleştirmek için sadeleştirir."""
    if not s:
        return ""
    s = str(s).strip().upper()
    s = s.replace("İ", "I").replace("I", "I").replace("Ş", "S").replace("Ğ", "G")
    s = s.replace("Ü", "U").replace("Ö", "O").replace("Ç", "C")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z]", "", s)
    ESLE = {
        "EYUP": "EYUPSULTAN",
        "EYUPSULTAN": "EYUPSULTAN",
        "GAZIOSMANPASA": "GAZIOSMANPASA",
    }
    return ESLE.get(s, s)


def to_xy(lon, lat, lat0):
    """Yerel düzlem projeksiyonu (km). Küçük alanlarda yeterli."""
    return (
        math.radians(lon) * R * math.cos(math.radians(lat0)),
        math.radians(lat) * R,
    )


def haversine(lon1, lat1, lon2, lat2):
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = p2 - p1
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def seg_dist(px, py, ax, ay, bx, by):
    dx, dy = bx - ax, by - ay
    L = dx * dx + dy * dy
    t = 0 if L == 0 else max(0, min(1, ((px - ax) * dx + (py - ay) * dy) / L))
    return math.hypot(px - (ax + t * dx), py - (ay + t * dy))


def dist_to_line(lon, lat, coords):
    """Noktanın LineString'e en kısa mesafesi (km)."""
    px, py = to_xy(lon, lat, lat)
    best, prev = float("inf"), None
    for c in coords:
        cur = to_xy(c[0], c[1], lat)
        if prev is not None:
            best = min(best, seg_dist(px, py, prev[0], prev[1], cur[0], cur[1]))
        prev = cur
    return best


def point_in_ring(lon, lat, ring):
    inside = False
    n = len(ring)
    for i in range(n):
        x1, y1 = ring[i][0], ring[i][1]
        x2, y2 = ring[(i + 1) % n][0], ring[(i + 1) % n][1]
        if (y1 > lat) != (y2 > lat):
            xint = (x2 - x1) * (lat - y1) / (y2 - y1) + x1
            if lon < xint:
                inside = not inside
    return inside


# --------------------------------------------------------------------------
# 1. ilçe sınırları (Overpass relation -> dış halka)
# --------------------------------------------------------------------------
def load_ilceler():
    raw = json.loads((DATA / "facilities" / "ilce_sinirlari_raw.json").read_text())
    ilceler = {}
    for el in raw["elements"]:
        ad = el["tags"].get("name")
        if not ad:
            continue
        # relation üyelerinin outer way'lerini birleştir
        parcalar = []
        for m in el.get("members", []):
            if m.get("role") == "outer" and m.get("geometry"):
                parcalar.append([(p["lon"], p["lat"]) for p in m["geometry"]])
        if not parcalar:
            continue
        halkalar = birlestir_halkalar(parcalar)
        lons = [p[0] for h in halkalar for p in h]
        lats = [p[1] for h in halkalar for p in h]
        ilceler[norm_ad(ad)] = {
            "ad": ad,
            "halkalar": halkalar,
            "bbox": (min(lons), min(lats), max(lons), max(lats)),
            "merkez": (sum(lons) / len(lons), sum(lats) / len(lats)),
        }
    return ilceler


def birlestir_halkalar(parcalar, tol=1e-7):
    """Kopuk way parçalarını uç uca ekleyerek kapalı halkalara çevirir."""
    kalan = [list(p) for p in parcalar]
    halkalar = []
    while kalan:
        cur = kalan.pop(0)
        degisti = True
        while degisti and abs(cur[0][0] - cur[-1][0]) + abs(cur[0][1] - cur[-1][1]) > tol:
            degisti = False
            for i, p in enumerate(kalan):
                if abs(cur[-1][0] - p[0][0]) + abs(cur[-1][1] - p[0][1]) < tol:
                    cur += p[1:]
                elif abs(cur[-1][0] - p[-1][0]) + abs(cur[-1][1] - p[-1][1]) < tol:
                    cur += p[::-1][1:]
                elif abs(cur[0][0] - p[-1][0]) + abs(cur[0][1] - p[-1][1]) < tol:
                    cur = p[:-1] + cur
                elif abs(cur[0][0] - p[0][0]) + abs(cur[0][1] - p[0][1]) < tol:
                    cur = p[::-1][:-1] + cur
                else:
                    continue
                kalan.pop(i)
                degisti = True
                break
        halkalar.append(cur)
    return halkalar


def ilce_bul(lon, lat, ilceler):
    for key, il in ilceler.items():
        x0, y0, x1, y1 = il["bbox"]
        if not (x0 <= lon <= x1 and y0 <= lat <= y1):
            continue
        if any(point_in_ring(lon, lat, h) for h in il["halkalar"]):
            return key
    return None


# --------------------------------------------------------------------------
# 2. tesisler: OSM + İBB -> tek normalize CSV
# --------------------------------------------------------------------------
TIP_ETIKET = {
    "hospital": "hastane",
    "fire_station": "itfaiye",
    "police": "polis",
    "assembly_point": "toplanma_alani",
}


def load_tesisler(ilceler):
    tesisler = []

    # --- OSM ---
    osm = json.loads((DATA / "facilities" / "osm_ist_raw.json").read_text())
    for el in osm["elements"]:
        t = el.get("tags", {})
        tip = TIP_ETIKET.get(t.get("amenity")) or TIP_ETIKET.get(t.get("emergency"))
        if not tip:
            continue
        lat = el.get("lat") or (el.get("center") or {}).get("lat")
        lon = el.get("lon") or (el.get("center") or {}).get("lon")
        if lat is None or lon is None:
            continue
        tesisler.append({
            "tip": tip,
            "ad": t.get("name", ""),
            "lat": round(lat, 6),
            "lon": round(lon, 6),
            "ilce": ilce_bul(lon, lat, ilceler) or "",
            "kaynak": "OSM",
        })

    # --- İBB itfaiye istasyonları (resmi) ---
    wb = openpyxl.load_workbook(DATA / "ibb" / "itfaiye_2025.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    for r in rows:
        if not r or r[4] is None or r[5] is None:
            continue
        try:
            lon, lat = float(r[4]), float(r[5])
        except (TypeError, ValueError):
            continue
        tesisler.append({
            "tip": "itfaiye",
            "ad": str(r[0] or ""),
            "lat": round(lat, 6),
            "lon": round(lon, 6),
            "ilce": norm_ad(r[1]),
            "kaynak": "IBB",
        })

    # --- İBB sağlık tesisleri (resmi) — sadece gerçek hastaneler ---
    wb = openpyxl.load_workbook(DATA / "ibb" / "saglik_tesisleri.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    for r in rows:
        if not r:
            continue
        alt = str(r[2] or "")
        if "Hastane" not in alt:
            continue
        try:
            lat, lon = float(r[6]), float(r[7])
        except (TypeError, ValueError):
            continue
        tesisler.append({
            "tip": "hastane",
            "ad": str(r[0] or ""),
            "lat": round(lat, 6),
            "lon": round(lon, 6),
            "ilce": norm_ad(r[3]),
            "kaynak": "IBB",
        })

    return tesisler


# --------------------------------------------------------------------------
# 3. Vs30 grid
# --------------------------------------------------------------------------
def load_vs30():
    pts = []
    with open(DATA / "vs30" / "istanbul_vs30.csv", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            pts.append((float(r["lat"]), float(r["lon"]), float(r["vs30"])))
    return pts


# --------------------------------------------------------------------------
# 4. İBB deprem senaryosu (mahalle bazlı, cp1254)
# --------------------------------------------------------------------------
SENARYO_SAYISAL = [
    "cok_agir_hasarli_bina_sayisi", "agir_hasarli_bina_sayisi",
    "orta_hasarli_bina_sayisi", "hafif_hasarli_bina_sayisi",
    "can_kaybi_sayisi", "agir_yarali_sayisi", "hastanede_tedavi_sayisi",
    "hafif_yarali_sayisi", "dogalgaz_boru_hasari", "icme_suyu_boru_hasari",
    "atik_su_boru_hasari", "gecici_barinma",
]


def load_senaryo():
    p = DATA / "ibb" / "deprem_senaryosu.csv"
    with open(p, encoding="cp1254") as fh:
        rows = list(csv.DictReader(fh, delimiter=";"))
    for r in rows:
        for k in SENARYO_SAYISAL:
            try:
                r[k] = int(str(r[k]).strip() or 0)
            except ValueError:
                r[k] = 0
    return rows


# --------------------------------------------------------------------------
# 5. İBB kentsel açık/yeşil alanlar — toplanma alanı proxy'si
# --------------------------------------------------------------------------
def poligon_alan_ha(halka, lat0):
    """Shoelace ile alan (hektar). Yerel düzlem projeksiyonu."""
    pts = [to_xy(p[0], p[1], lat0) for p in halka]
    s = 0.0
    for i in range(len(pts)):
        x1, y1 = pts[i]
        x2, y2 = pts[(i + 1) % len(pts)]
        s += x1 * y2 - x2 * y1
    return abs(s) / 2 * 100  # km² -> hektar


def load_yesil_alanlar():
    gj = json.loads((DATA / "ibb" / "yesil_alanlar.geojson").read_text())
    alanlar = []
    for f in gj["features"]:
        g, p = f["geometry"], f["properties"]
        if g["type"] == "Polygon":
            halkalar = [g["coordinates"][0]]
        elif g["type"] == "MultiPolygon":
            halkalar = [poly[0] for poly in g["coordinates"]]
        else:
            continue
        tum = [pt for h in halkalar for pt in h]
        if not tum:
            continue
        lat0 = sum(pt[1] for pt in tum) / len(tum)
        alanlar.append({
            "ad": p.get("MAHALLE", ""),
            "tur": p.get("TUR", ""),
            "ilce": norm_ad(p.get("ILCE")),
            "lat": round(lat0, 6),
            "lon": round(sum(pt[0] for pt in tum) / len(tum), 6),
            "alan_ha": round(sum(poligon_alan_ha(h, lat0) for h in halkalar), 3),
        })
    return alanlar


# --------------------------------------------------------------------------
# 6. Grid bazlı erişim analizi — her Vs30 hücresinden en yakın tesise mesafe
# --------------------------------------------------------------------------
def erisim_gridi(vs30, tesisler, ilceler):
    import numpy as np

    grid = np.array([(la, lo, v) for (la, lo, v) in vs30])
    glat, glon = grid[:, 0], grid[:, 1]
    lat0 = float(glat.mean())
    kx = math.radians(1) * R * math.cos(math.radians(lat0))
    ky = math.radians(1) * R
    gx, gy = glon * kx, glat * ky

    sonuc = {"lat": glat, "lon": glon, "vs30": grid[:, 2]}
    for tip in ("hastane", "itfaiye", "polis", "toplanma_alani"):
        pts = [(t["lon"], t["lat"]) for t in tesisler if t["tip"] == tip]
        if not pts:
            sonuc[tip] = np.full(len(grid), np.nan)
            continue
        arr = np.array(pts)
        tx, ty = arr[:, 0] * kx, arr[:, 1] * ky
        # chunk'layarak mesafe matrisi (bellek için)
        en_yakin = np.full(len(grid), np.inf)
        for i in range(0, len(grid), 4000):
            dx = gx[i:i + 4000, None] - tx[None, :]
            dy = gy[i:i + 4000, None] - ty[None, :]
            en_yakin[i:i + 4000] = np.sqrt(dx * dx + dy * dy).min(axis=1)
        sonuc[tip] = en_yakin

    # her hücreyi ilçeye ata (bbox ön eleme + point-in-polygon)
    ilce_ad = []
    for la, lo in zip(glat, glon):
        k = ilce_bul(lo, la, ilceler)
        ilce_ad.append(ilceler[k]["ad"] if k else "")
    sonuc["ilce"] = ilce_ad
    return sonuc


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main():
    print("ilçe sınırları okunuyor...")
    ilceler = load_ilceler()
    print(f"  {len(ilceler)} ilçe")

    print("tesisler normalize ediliyor...")
    tesisler = load_tesisler(ilceler)
    with open(DERIVED / "tesisler.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["tip", "ad", "lat", "lon", "ilce", "kaynak"])
        w.writeheader()
        w.writerows(tesisler)
    say = defaultdict(int)
    for t in tesisler:
        say[t["tip"]] += 1
    print("  ", dict(say))

    print("senaryo okunuyor...")
    senaryo = load_senaryo()
    with open(DERIVED / "mahalle_senaryo.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(senaryo[0].keys()))
        w.writeheader()
        w.writerows(senaryo)
    print(f"  {len(senaryo)} mahalle")

    print("Vs30 okunuyor...")
    vs30 = load_vs30()

    print("faylar okunuyor...")
    faylar = json.loads((DATA / "faults" / "tr_faults_imp.geojson").read_text())["features"]

    print("depremler okunuyor...")
    eq = json.loads((DATA / "earthquakes" / "ist_eq_all.geojson").read_text())["features"]
    with open(DERIVED / "depremler.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["tarih_utc", "mag", "derinlik_km", "lat", "lon", "yer", "usgs_id"])
        import datetime as dt
        for x in eq:
            p, c = x["properties"], x["geometry"]["coordinates"]
            ts = dt.datetime.fromtimestamp(p["time"] / 1000, dt.UTC)
            w.writerow([ts.strftime("%Y-%m-%d %H:%M:%S"), p.get("mag"), c[2],
                        round(c[1], 4), round(c[0], 4), p.get("place", ""), x.get("id", "")])
    print(f"  {len(eq)} deprem")

    # --- ilçe özet tablosu ---
    print("ilçe özeti hesaplanıyor...")
    senaryo_ilce = defaultdict(lambda: defaultdict(int))
    for r in senaryo:
        k = norm_ad(r["ilce_adi"])
        for c in SENARYO_SAYISAL:
            senaryo_ilce[k][c] += r[c]

    tesis_ilce = defaultdict(lambda: defaultdict(list))
    for t in tesisler:
        if t["ilce"]:
            tesis_ilce[t["ilce"]][t["tip"]].append(t)

    print("yeşil alanlar okunuyor...")
    yesil = load_yesil_alanlar()
    with open(DERIVED / "yesil_alanlar.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["ad", "tur", "ilce", "lat", "lon", "alan_ha"])
        w.writeheader()
        w.writerows(yesil)
    yesil_ilce = defaultdict(lambda: {"adet": 0, "ha": 0.0})
    for y in yesil:
        yesil_ilce[y["ilce"]]["adet"] += 1
        yesil_ilce[y["ilce"]]["ha"] += y["alan_ha"]
    print(f"  {len(yesil)} açık/yeşil alan, toplam "
          f"{sum(v['ha'] for v in yesil_ilce.values()):.0f} ha")

    print("grid erişim analizi...")
    er = erisim_gridi(vs30, tesisler, ilceler)
    with open(DERIVED / "grid_erisim.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["lat", "lon", "vs30", "ilce", "hastane_km", "itfaiye_km",
                    "polis_km", "toplanma_km"])
        for i in range(len(er["lat"])):
            w.writerow([round(er["lat"][i], 6), round(er["lon"][i], 6),
                        round(float(er["vs30"][i]), 1), er["ilce"][i],
                        round(float(er["hastane"][i]), 2),
                        round(float(er["itfaiye"][i]), 2),
                        round(float(er["polis"][i]), 2),
                        round(float(er["toplanma_alani"][i]), 2)])
    ic_idx = [i for i in range(len(er["lat"])) if er["ilce"][i]]
    print(f"  {len(er['lat'])} hücre ({len(ic_idx)} tanesi İstanbul sınırı içinde)")

    # ilçe içi grid ortalamaları (erişim boşluğu)
    grid_ilce = defaultdict(lambda: defaultdict(list))
    for i in ic_idx:
        k = norm_ad(er["ilce"][i])
        for tip in ("hastane", "itfaiye", "polis", "toplanma_alani"):
            grid_ilce[k][tip].append(float(er[tip][i]))

    ozet = []
    for key, il in sorted(ilceler.items()):
        lon, lat = il["merkez"]

        # Vs30: merkeze en yakın grid hücresi + ilçe bbox içi ortalama
        x0, y0, x1, y1 = il["bbox"]
        ic = [v for (la, lo, v) in vs30 if x0 <= lo <= x1 and y0 <= la <= y1]
        en_yakin = min(vs30, key=lambda p: (p[0] - lat) ** 2 + (p[1] - lon) ** 2)[2]

        # en yakın fay
        fay_km, fay_ad = float("inf"), ""
        for f in faylar:
            d = dist_to_line(lon, lat, f["geometry"]["coordinates"])
            if d < fay_km:
                fay_km = d
                p = f["properties"]
                fay_ad = (p.get("FAULT_NAME") or p.get("ZONE_NAME") or "?").strip()

        # merkeze en yakın tesis mesafeleri (ilçe sınırı gözetmeksizin, tüm il)
        def en_yakin_tesis(tip):
            aday = [t for t in tesisler if t["tip"] == tip]
            if not aday:
                return None, ""
            best = min(aday, key=lambda t: haversine(lon, lat, t["lon"], t["lat"]))
            return round(haversine(lon, lat, best["lon"], best["lat"]), 2), best["ad"]

        h_km, h_ad = en_yakin_tesis("hastane")
        i_km, i_ad = en_yakin_tesis("itfaiye")
        p_km, _ = en_yakin_tesis("polis")
        t_km, _ = en_yakin_tesis("toplanma_alani")

        s = senaryo_ilce.get(key, {})
        ti = tesis_ilce.get(key, {})
        ya = yesil_ilce.get(key, {"adet": 0, "ha": 0.0})
        gi = grid_ilce.get(key, {})

        def ort(tip):
            v = gi.get(tip, [])
            return round(sum(v) / len(v), 2) if v else ""

        def enkotu(tip):
            v = gi.get(tip, [])
            return round(max(v), 2) if v else ""

        barinma = s.get("gecici_barinma", 0)
        ozet.append({
            "ilce": il["ad"],
            "vs30_merkez": round(en_yakin, 1),
            "vs30_ort": round(sum(ic) / len(ic), 1) if ic else "",
            "vs30_min": round(min(ic), 1) if ic else "",
            "en_yakin_fay": fay_ad,
            "fay_km": round(fay_km, 1),
            "senaryo_can_kaybi": s.get("can_kaybi_sayisi", ""),
            "senaryo_agir_hasarli_bina": (s.get("cok_agir_hasarli_bina_sayisi", 0)
                                          + s.get("agir_hasarli_bina_sayisi", 0)) if s else "",
            "senaryo_gecici_barinma": barinma,
            "hastane_sayisi": len(ti.get("hastane", [])),
            "itfaiye_sayisi": len(ti.get("itfaiye", [])),
            "polis_sayisi": len(ti.get("polis", [])),
            "toplanma_alani_sayisi": len(ti.get("toplanma_alani", [])),
            "acik_yesil_alan_adet": ya["adet"],
            "acik_yesil_alan_ha": round(ya["ha"], 1),
            # barınma ihtiyacı başına düşen açık alan (m²/kişi)
            "barinma_basina_m2": (round(ya["ha"] * 10000 / barinma, 1)
                                  if barinma else ""),
            "en_yakin_hastane_km": h_km,
            "en_yakin_itfaiye_km": i_km,
            "en_yakin_polis_km": p_km,
            "en_yakin_toplanma_km": t_km,
            # ilçe alanı geneline yayılmış erişim (grid ortalaması / en kötü hücre)
            "grid_ort_hastane_km": ort("hastane"),
            "grid_ort_itfaiye_km": ort("itfaiye"),
            "grid_enkotu_itfaiye_km": enkotu("itfaiye"),
            "grid_ort_toplanma_km": ort("toplanma_alani"),
        })

    ozet.sort(key=lambda r: -(r["senaryo_can_kaybi"] or 0))
    with open(DERIVED / "istanbul_ilce_ozet.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(ozet[0].keys()))
        w.writeheader()
        w.writerows(ozet)
    print(f"  {len(ozet)} ilçe -> derived/istanbul_ilce_ozet.csv")


if __name__ == "__main__":
    main()
